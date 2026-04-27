from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.augment_binary_interpretation_matrix import SCBE_EXTENSION_SHEETS, augment_workbook


def test_augment_binary_interpretation_matrix_adds_five_sheets(tmp_path: Path) -> None:
    workbook = tmp_path / "binary.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Cover"
    ws.append(["Binary Interpretation Matrix"])
    wb.save(workbook)

    report = augment_workbook(workbook)
    reloaded = load_workbook(workbook, read_only=True, data_only=False)

    assert report["openxml_load_check"] == "PASSED"
    assert report["sheet_count"] == 6
    assert Path(report["backup_path"]).exists()
    for sheet in SCBE_EXTENSION_SHEETS:
        assert sheet in reloaded.sheetnames
        assert report["extension_sheets"][sheet] is True

    periodic = reloaded["Periodic Table Mapping"]
    assert periodic.max_row == 65
    assert periodic["A1"].value == "Bucket"
    assert periodic["L2"].value.startswith("byte_to_element_index")

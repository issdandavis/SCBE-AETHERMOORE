#!/usr/bin/env python3
"""Convert binary_interpretation_matrix.xlsx (20-sheet Kimi feedback) to SFT.

Each sheet becomes a set of lookup/teach records:
  - sheet header → "What is the binary representation of X?" style prompts
  - row-level → "value -> binary/hex/desc" instruction/output pairs
  - cross-sheet → "How does <concept> appear in <DNA/periodic/quantum>?" pairs

Output: training-data/sft/binary_matrix_v2_full.sft.jsonl
Format: messages-style ChatML for SFTTrainer.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
SRC = Path(r"C:\Users\issda\Downloads\binary_interpretation_matrix.xlsx")
OUT = REPO / "training-data" / "sft" / "binary_matrix_v2_full.sft.jsonl"

SYSTEM = (
    "You are a binary/agentic-coding teaching assistant. You translate fluently "
    "between binary, decimal, hexadecimal, ASCII, IEEE 754, biological encodings, "
    "and quantum representations. Answer with the canonical mapping or operation."
)


def clean(c) -> str:
    if c is None:
        return ""
    s = str(c).strip()
    return s.replace("→", "->").replace("—", "-").replace("–", "-")


def make(prompt: str, answer: str, sheet: str) -> dict:
    return {
        "messages": [
            {"role": "system", "content": SYSTEM},
            {"role": "user", "content": prompt.strip()},
            {"role": "assistant", "content": answer.strip()},
        ],
        "metadata": {"source": "binary_matrix_v2_xlsx", "sheet": sheet},
    }


def rows(ws):
    out = []
    for row in ws.iter_rows(values_only=True):
        cells = [clean(c) for c in row]
        if any(cells):
            out.append(cells)
    return out


def first_data_row(rs):
    """Find the first row that looks like a header (>=2 non-empty cells)."""
    for i, r in enumerate(rs):
        non_empty = [c for c in r if c]
        if len(non_empty) >= 3 and any(any(ch.isalpha() for ch in c) for c in r):
            return i
    return 0


def harvest_table(rs, sheet_name: str) -> list[dict]:
    """Generic harvester: treat first qualifying row as header, then each row as a record."""
    if not rs:
        return []
    hdr_idx = first_data_row(rs)
    header = [c for c in rs[hdr_idx]]
    records = []
    for r in rs[hdr_idx + 1 :]:
        if not any(r):
            continue
        non_empty_cols = sum(1 for c in r if c)
        if non_empty_cols < 2:
            continue
        cells = list(zip(header, r))
        # First non-empty cell = key
        key = next((v for h, v in cells if v), None)
        if not key:
            continue
        # Compose answer from remaining
        parts = [f"{h}: {v}" for h, v in cells if h and v and h != header[0]]
        if not parts:
            continue
        prompt = (
            f"In the {sheet_name} reference, what is the full mapping for `{key}`?"
        )
        answer = "\n".join(parts)
        records.append(make(prompt, answer, sheet_name))
    return records


def harvest_intro(rs, sheet_name: str) -> dict | None:
    """Pull the title + description from top of sheet as a single 'overview' record."""
    title = None
    desc = None
    for r in rs[:6]:
        for c in r:
            if not c:
                continue
            if title is None and len(c) > 5 and c[0].isupper():
                title = c
                continue
            if title and desc is None and len(c) > 20:
                desc = c
                break
        if title and desc:
            break
    if not title or not desc:
        return None
    prompt = f"What does the '{sheet_name}' sheet of the Binary Interpretation Matrix cover?"
    answer = f"{title}\n{desc}"
    return make(prompt, answer, sheet_name)


def main():
    if not SRC.exists():
        print(f"ERROR: {SRC} not found", file=sys.stderr)
        return 2
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    records: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rs = rows(ws)
        intro = harvest_intro(rs, sheet_name)
        if intro:
            records.append(intro)
        records.extend(harvest_table(rs, sheet_name))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as fh:
        for r in records:
            fh.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"wrote {len(records)} records -> {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

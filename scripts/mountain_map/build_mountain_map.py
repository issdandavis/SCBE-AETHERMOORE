"""Build the cross-language "mountain map" from the verified construct data.

Three views across 18 language faces. NOTE: this is a NOTATION map, not a computation map --
view 2 is a syntactic phylogeny (surface spelling), which anti-correlates with semantics; the
same-computation axis is the IR + polyglot_conformance.py, not this.

  1. construct_table.csv / .xlsx -- the basic table: every curated construct x every
     language, with a per-cell confidence (the map carries its own provenance).
  2. mountain.dot / mountain.md  -- a SYNTACTIC PHYLOGENY: languages are nodes, edge weight
     is how many constructs two languages spell IDENTICALLY (computed from the table, not
     asserted). It is a family tree of NOTATION (brace / colon / ML family), NOT computational
     distance -- surface spelling anti-correlates with semantics (`==` is one glyph with three
     meanings; `map f xs` vs a comprehension is one meaning in two spellings). Use it as a
     transpiler work-allocation map; the same-computation axis is the IR + polyglot_conformance.py
     (which RUNS the backends). See semantic_vs_syntax.py for the executed proof.
  3. pipeline_map.md             -- what happens when you hit Enter: per language, the
     source -> lex -> parse -> IR -> execute path.

It also CROSS-CHECKS itself against the older hand-made python/scbe/cross_lang.py table and
reports any cell where the two disagree -- so the map can catch its own errors.

This maps the shared CONSTRUCTS and where the faces diverge. It is NOT "every command":
the standard library is unbounded; that ceiling never closes. The keyword/operator skeleton
is the easy, finite part -- the value here is the verified divergences.

    python scripts/mountain_map/build_mountain_map.py            # write all artifacts
    python scripts/mountain_map/build_mountain_map.py --summary  # just print the summary
"""

from __future__ import annotations

import csv
import json
import sys
from itertools import combinations
from pathlib import Path
from typing import Dict, List

HERE = Path(__file__).resolve().parent
REPO = HERE.parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

DATA = HERE / "mountain_map_data.json"


def _norm(code: str) -> str:
    """Lenient construct identity: ignore whitespace, case, trailing ';'."""
    return "".join(code.split()).rstrip(";").lower()


def load() -> dict:
    return json.loads(DATA.read_text(encoding="utf-8"))


def _flat(code: str) -> str:
    return code.replace("\n", " ⏎ ")


def write_table_csv(data: dict, out: Path) -> Path:
    langs: List[str] = data["meta"]["languages"]
    grid = data["grid"]
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["construct"] + langs)
        for k in data["meta"]["constructs"]:
            w.writerow([k] + [_flat(grid[k][lang]["code"]) for lang in langs])
    return out


def write_table_xlsx(data: dict, out: Path) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return None
    langs = data["meta"]["languages"]
    grid = data["grid"]
    fills = {
        "high": PatternFill("solid", fgColor="E6F4EA"),
        "medium": PatternFill("solid", fgColor="FEF7E0"),
        "low": PatternFill("solid", fgColor="FCE8E6"),
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "constructs"
    ws.append(["construct"] + langs)
    for c in ws[1]:
        c.font = Font(bold=True)
    for k in data["meta"]["constructs"]:
        ws.append([k] + [_flat(grid[k][lang]["code"]) for lang in langs])
        row = ws.max_row
        ws.cell(row, 1).font = Font(bold=True)
        for ci, lang in enumerate(langs, start=2):
            ws.cell(row, ci).fill = fills.get(grid[k][lang]["confidence"], fills["high"])
    ws.freeze_panes = "B2"
    ws2 = wb.create_sheet("pipeline")
    cols = ["language", "exec_model", "lex", "parse", "ir", "runtime", "on_enter"]
    ws2.append(cols)
    for c in ws2[1]:
        c.font = Font(bold=True)
    for lang in langs:
        p = data["pipeline"][lang]
        ws2.append([lang] + [p[k] for k in cols[1:]])
    for ws_ in (ws, ws2):
        for col in ws_.columns:
            ws_.column_dimensions[col[0].column_letter].width = 22
            for cell in col:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(out)
    return out


def similarity(data: dict) -> Dict[str, Dict[str, float]]:
    """Pairwise: fraction of constructs two languages spell identically (normalized)."""
    langs = data["meta"]["languages"]
    grid = data["grid"]
    constructs = data["meta"]["constructs"]
    sim: Dict[str, Dict[str, float]] = {a: {} for a in langs}
    for a, b in combinations(langs, 2):
        same = sum(1 for k in constructs if _norm(grid[k][a]["code"]) == _norm(grid[k][b]["code"]))
        s = round(same / len(constructs), 3)
        sim[a][b] = sim[b][a] = s
    return sim


def write_mountain(data: dict, sim: Dict[str, Dict[str, float]], dot: Path, md: Path, threshold: float = 0.4) -> None:
    langs = data["meta"]["languages"]
    lines = [
        "graph mountain {",
        "  layout=neato;",
        "  overlap=false;",
        "  splines=true;",
        '  node [shape=box, style="rounded,filled", fillcolor="#eef3fb", fontname="Helvetica"];',
    ]
    for a, b in combinations(langs, 2):
        s = sim[a][b]
        if s >= threshold:
            lines.append('  "%s" -- "%s" [label="%.2f", penwidth=%.2f];' % (a, b, s, 0.5 + 5 * s))
    lines.append("}")
    dot.write_text("\n".join(lines) + "\n", encoding="utf-8")

    rows = [
        "# Syntactic phylogeny: language similarity by identical construct spelling",
        "",
        "Edge = fraction of the %d constructs two languages spell *identically* (normalized)."
        % len(data["meta"]["constructs"]),
        "Computed from the table, not asserted -- but this is SURFACE spelling, a family tree of",
        "NOTATION (brace / colon / ML family), NOT the shape of the computation.",
        "",
        "## What this is NOT",
        "",
        "Surface spelling is not semantics, and the two come apart:",
        "- identical spelling, different meaning: `==` is one glyph in Java/C#/JS/Python but means",
        '  reference vs value vs coercing equality (`1 == "1"` is False in Python, true in JS).',
        "- different spelling, same computation: Haskell `map f xs` and Python `[f(x) for x in xs]`",
        "  compute the same thing -- yet Haskell scores as the outlier here.",
        "So these distances are notation lineage, not semantic distance. Use this as a TRANSPILER",
        "WORK-ALLOCATION map: identical cells emit trivially; the divergences (the Haskell column)",
        "are where real semantic effort lives. The same-computation axis is the IR +",
        "polyglot_conformance.py (which RUNS the backends); see semantic_vs_syntax.py for proof.",
        "",
        "## Nearest neighbours (each language's 3 closest faces by notation)",
        "",
    ]
    for a in langs:
        nbrs = sorted(((b, sim[a][b]) for b in langs if b != a), key=lambda x: x[1], reverse=True)[:3]
        rows.append("- **%s** -> %s" % (a, ", ".join("%s (%.2f)" % (b, s) for b, s in nbrs)))
    md.write_text("\n".join(rows) + "\n", encoding="utf-8")


def write_pipeline(data: dict, out: Path) -> Path:
    langs = data["meta"]["languages"]
    rows = [
        "# What happens when you hit Enter: the source -> execute pipeline",
        "",
        "| language | model | lex | parse | IR | runtime | on Enter |",
        "|---|---|---|---|---|---|---|",
    ]
    for lang in langs:
        p = data["pipeline"][lang]
        cells = [lang, p["exec_model"], p["lex"], p["parse"], p["ir"], p["runtime"], p["on_enter"]]
        rows.append("| " + " | ".join(c.replace("|", "\\|") for c in cells) + " |")
    out.write_text("\n".join(rows) + "\n", encoding="utf-8")
    return out


# old hand-made concept -> our construct key (overlap for the self-check)
_XCHECK = {
    "print": "print_stdout",
    "function": "func_def",
    "for_loop": "for_range",
    "if": "if_else",
    "variable": "var_mutable",
    "list": "list_literal",
    "length": "length",
    "comment": "line_comment",
    "true": "bool_literal",
    "null": "null_literal",
    "return": "return",
    "while": "while",
    "string": "string_literal",
    "import": "import_module",
}


def cross_check(data: dict) -> dict:
    """Compare overlapping cells against the older python/scbe/cross_lang.py ROSETTA table."""
    try:
        from python.scbe.cross_lang import ROSETTA
    except Exception as e:
        return {"error": str(e)}
    grid = data["grid"]
    agree, conflicts = 0, []
    for concept, key in _XCHECK.items():
        row = ROSETTA.get(concept, {})
        for lang, old in row.items():
            new = grid.get(key, {}).get(lang)
            if not new:
                continue
            if _norm(old) == _norm(new["code"]):
                agree += 1
            else:
                conflicts.append({"construct": key, "language": lang, "old": old, "new": new["code"]})
    return {"agree": agree, "conflicts": conflicts}


def build(summary_only: bool = False) -> dict:
    data = load()
    sim = similarity(data)
    xc = cross_check(data)
    if not summary_only:
        write_table_csv(data, HERE / "construct_table.csv")
        write_table_xlsx(data, HERE / "construct_table.xlsx")
        write_mountain(data, sim, HERE / "mountain.dot", HERE / "mountain.md")
        write_pipeline(data, HERE / "pipeline_map.md")
    # a few honest, deterministic facts for the summary
    cluster_pairs = sorted(((a, b, sim[a][b]) for a in sim for b in sim[a] if a < b), key=lambda x: x[2], reverse=True)[
        :5
    ]
    return {"data": data, "sim": sim, "xcheck": xc, "top_pairs": cluster_pairs}


def main(argv: List[str] | None = None) -> int:
    summary_only = "--summary" in (argv if argv is not None else sys.argv[1:])
    r = build(summary_only=summary_only)
    d = r["data"]
    n_cells = sum(len(d["grid"][k]) for k in d["meta"]["constructs"])
    print(
        "MOUNTAIN MAP  %d constructs x %d languages = %d cells"
        % (len(d["meta"]["constructs"]), len(d["meta"]["languages"]), n_cells)
    )
    if not summary_only:
        print("  wrote: construct_table.csv, construct_table.xlsx, mountain.dot, mountain.md, pipeline_map.md")
    print("  closest faces:", ", ".join("%s~%s %.2f" % (a, b, s) for a, b, s in r["top_pairs"]))
    xc = r["xcheck"]
    if "error" in xc:
        print("  cross-check vs cross_lang.py: unavailable (%s)" % xc["error"])
    else:
        print("  cross-check vs cross_lang.py: %d agree, %d conflicts" % (xc["agree"], len(xc["conflicts"])))
        for c in xc["conflicts"][:8]:
            print("    CONFLICT %s/%s: old %r vs new %r" % (c["construct"], c["language"], c["old"], c["new"]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

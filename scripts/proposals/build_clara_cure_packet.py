from __future__ import annotations

import json
import os
import shutil
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


ROOT = Path(__file__).resolve().parents[2]
PROPOSALS = ROOT / "docs" / "proposals"
CURE = PROPOSALS / "DARPA_CLARA_CURE"
OUT = CURE / "generated"


VOLUME_1 = PROPOSALS / "CLARA_Volume1_Technical.pdf"
VOLUME_3 = PROPOSALS / "CLARA_Volume3_Administrative.pdf"
SUMMARY_SLIDE = OUT / "CLARA_Summary_Slide.pdf"
PRICE_WORKBOOK = OUT / "CLARA_Volume2_Price_Workbook.xlsx"
PACKAGE_ZIP = OUT / "CLARA_FP033_corrected_packet_local.zip"
CHECK_JSON = OUT / "submission_data_check.json"


MILESTONES = [
    ("M1", "Phase 1", "Standardized benchmark integration and reproducible runner scripts", 160000),
    ("M2", "Phase 1", "Invariant artifact export, trace schema, expanded tests", 200000),
    ("M3", "Phase 1", "Defeasible rule export and bounded-unfolding explanation traces", 210000),
    ("M4", "Phase 1", "Hackathon readiness package, container, deterministic logs", 200000),
    ("M5", "Phase 2", "Training-time adaptation with artifact preservation and rollback", 295000),
    ("M6", "Phase 2", "Multi-domain adaptation and sample-complexity reporting", 295000),
    ("M7", "Phase 2", "Hackathon iteration, threshold tuning, trace interpretability", 295000),
]


def _money(value: int) -> str:
    return f"${value:,.0f}"


def build_summary_slide() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(SUMMARY_SLIDE),
        pagesize=landscape(letter),
        rightMargin=0.35 * inch,
        leftMargin=0.35 * inch,
        topMargin=0.3 * inch,
        bottomMargin=0.25 * inch,
    )
    title = ParagraphStyle(
        "title",
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=20,
        textColor=colors.HexColor("#111827"),
    )
    h = ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=9, leading=11, textColor=colors.HexColor("#1f2937"))
    p = ParagraphStyle("p", fontName="Helvetica", fontSize=7.6, leading=9.2, textColor=colors.HexColor("#111827"))
    small = ParagraphStyle("small", fontName="Helvetica", fontSize=7.2, leading=8.7, textColor=colors.HexColor("#111827"))

    def block(label: str, body: str):
        return [Paragraph(label, h), Spacer(1, 0.025 * inch), Paragraph(body, p)]

    meta = (
        "Solicitation: DARPA-PA-25-07-02 CLARA | TA1 Performer | PI: Issac Daniel Davis | "
        "Org: AetherMoore / Issac D. Davis | UEI: J4NXHM6N5F59 | Draft cost: $1.655M"
    )
    left = []
    left += block(
        "Problem",
        "Current AI safety usually attaches automated reasoning around ML after the fact. "
        "That catches known attacks but does not provide compositional assurance, proof-carrying "
        "explanations, or reliable behavior under novel attack forms.",
    )
    left += [Spacer(1, 0.08 * inch)]
    left += block(
        "Technical Idea",
        "SCBE-AETHERMOORE places automated reasoning inside the inference pipeline through "
        "defeasible LP governance, concept-bottleneck state, geometric intent verification, "
        "typed composition layers, and machine-checkable receipts.",
    )
    left += [Spacer(1, 0.08 * inch)]
    left += block(
        "CLARA Fit",
        "Targets tight ML+AR integration, automated logical proof, hierarchical explainability, "
        "compositional reuse, and tractable high-assurance behavior.",
    )

    mid = []
    mid += block(
        "Phase 1 Deliverables",
        "M1 benchmark runner; M2 invariant artifact export and trace schema; M3 defeasible-rule "
        "export with bounded unfolding; M4 hackathon readiness container and deterministic logs.",
    )
    mid += [Spacer(1, 0.08 * inch)]
    mid += block(
        "Phase 2 Deliverables",
        "Training-time adaptation with rollback, multi-domain adaptation, sample-complexity "
        "reporting, proof/export integration, and trace interpretability.",
    )
    mid += [Spacer(1, 0.08 * inch)]
    mid += block(
        "Evidence",
        "Local proposal materials cite 950+ automated tests, attack-chain evaluation automation, "
        "governance pipeline artifacts, concept-bottleneck state, and USPTO #63/961,403.",
    )

    right = []
    right += block(
        "Risks And Mitigations",
        "Proof export maturity: Phase 1 proof stubs and trace schema. LP toolchain integration: "
        "explicit defeasible-rule export and compatibility mapping. Evaluation transfer: "
        "deterministic runners and benchmark receipts.",
    )
    right += [Spacer(1, 0.08 * inch)]
    right += block(
        "Cost",
        "Phase 1: $770K over 15 months. Phase 2: $885K over 9 months. Total: $1.655M, under the "
        "$2M CLARA cap and under the $1.94M non-hackathon reference.",
    )
    right += [Spacer(1, 0.08 * inch)]
    right += block(
        "Correction Note",
        "Generated as the missing Summary Slide attachment for local cure packaging. Transfer to "
        "the official DARPA template if the Government requires the provided template format.",
    )

    story = [
        Paragraph("SCBE-AETHERMOORE: Verifiable Defeasible LP Composition for Trustworthy Multi-Agent AI Governance", title),
        Paragraph(meta, small),
        Spacer(1, 0.12 * inch),
        Table([[left, mid, right]], colWidths=[3.35 * inch, 3.35 * inch, 3.35 * inch]),
    ]
    doc.build(story)


def build_price_workbook() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Summary"
    ws.append(["Field", "Value"])
    rows = [
        ("Solicitation", "DARPA-PA-25-07-02 CLARA"),
        ("Proposal Ref", "DARPA-PA-25-07-02-CLARA-FP-033"),
        ("Title", "SCBE-AETHERMOORE: Verifiable Defeasible LP Composition for Trustworthy Multi-Agent AI Governance"),
        ("PI", "Issac Daniel Davis"),
        ("UEI", "J4NXHM6N5F59"),
        ("Phase 1 Base", 770000),
        ("Phase 2 Option", 885000),
        ("Total Proposed Cost", 1655000),
        ("Program Cap", 2000000),
        ("Non-Hackathon Reference", 1940000),
        ("Generated UTC", datetime.now(timezone.utc).isoformat()),
    ]
    for row in rows:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100

    ws2 = wb.create_sheet("Milestones")
    ws2.append(["Milestone", "Phase", "Deliverable", "Draft Amount"])
    for milestone in MILESTONES:
        ws2.append(list(milestone))
    ws2.append(["", "", "Total", f"=SUM(D2:D{len(MILESTONES)+1})"])
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 86
    ws2.column_dimensions["D"].width = 18

    ws3 = wb.create_sheet("Workbook Checklist")
    checks = [
        "Use the official TA1 Streamlined Cost Buildup Workbook if DARPA requires that exact attachment.",
        "Copy these totals and milestones into the official Schedule of Milestones and Payments tab.",
        "Match total proposed cost across portal entry, Summary Slide, Volume 1, Volume 2, and Volume 3.",
        "Confirm indirect/overhead basis before final submission.",
        "Confirm hackathon incentive/payment handling against the CLARA DO and FAQ.",
        "Attach this workbook only if DARPA accepts non-template cure materials; otherwise attach the official template.",
    ]
    ws3.append(["Check"])
    for check in checks:
        ws3.append([check])
    ws3.column_dimensions["A"].width = 120

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(PRICE_WORKBOOK)


def fetch_sam_gov_submission_data() -> dict:
    api_key = (os.getenv("SAM_GOV_API_KEY") or os.getenv("DATA_GOV_API_KEY") or "").strip()
    if not api_key:
        return {
            "attempted": False,
            "reason": "missing SAM_GOV_API_KEY or DATA_GOV_API_KEY in environment",
            "records": [],
        }

    params = {
        "api_key": api_key,
        "solnum": "DARPA-PA-25-07-02",
        "postedFrom": "01/01/2026",
        "postedTo": datetime.now(timezone.utc).strftime("%m/%d/%Y"),
        "limit": 10,
        "offset": 0,
    }
    url = "https://api.sam.gov/opportunities/v2/search?" + urlencode(params)
    request = Request(url, method="GET")
    try:
        with urlopen(request, timeout=20) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        return {
            "attempted": True,
            "ok": False,
            "http_status": exc.code,
            "error": exc.read().decode("utf-8", errors="replace")[:1000],
            "records": [],
        }
    except (URLError, TimeoutError) as exc:
        return {
            "attempted": True,
            "ok": False,
            "error": repr(exc),
            "records": [],
        }

    raw_records = payload.get("opportunitiesData") or payload.get("data") or payload.get("results") or []
    if isinstance(raw_records, dict):
        raw_records = [raw_records]

    records = []
    for item in raw_records[:10]:
        records.append(
            {
                "title": item.get("title") or item.get("opportunityTitle"),
                "solicitationNumber": item.get("solicitationNumber") or item.get("solNumber"),
                "noticeId": item.get("noticeId") or item.get("id"),
                "postedDate": item.get("postedDate"),
                "responseDeadline": item.get("responseDeadLine") or item.get("responseDeadline"),
                "active": item.get("active"),
                "uiLink": item.get("uiLink") or item.get("url"),
            }
        )

    return {
        "attempted": True,
        "ok": True,
        "record_count": len(records),
        "records": records,
    }


def build_check() -> dict:
    env_status = {
        name: bool(os.getenv(name, "").strip())
        for name in ("SAM_GOV_API_KEY", "DATA_GOV_API_KEY", "DARPA_API_KEY", "BAA_API_KEY")
    }
    required = {
        "summary_slide": SUMMARY_SLIDE,
        "volume1_technical": VOLUME_1,
        "volume2_price": PRICE_WORKBOOK,
        "volume3_administrative": VOLUME_3,
    }
    packet = {
        key: {
            "path": str(path),
            "exists": path.exists(),
            "bytes": path.stat().st_size if path.exists() else 0,
        }
        for key, path in required.items()
    }
    total = sum(amount for *_rest, amount in MILESTONES)
    data = {
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "proposal_ref": "DARPA-PA-25-07-02-CLARA-FP-033",
        "credential_status": env_status,
        "api_note": (
            "Local repo contains SAM.gov public opportunity API client and DARPA portal text parser, "
            "but no DARPA submission-status API client. GitHub repo secrets can provide the SAM.gov "
            "key during Actions runs, but secret values cannot be read back locally."
        ),
        "sam_gov_live_check": fetch_sam_gov_submission_data(),
        "packet": packet,
        "price": {
            "phase1": 770000,
            "phase2": 885000,
            "milestone_total": total,
            "declared_total": 1655000,
            "matches_declared_total": total == 1655000,
            "under_2m_cap": total <= 2000000,
            "under_1940k_non_hackathon_reference": total <= 1940000,
        },
        "nonconformance_root_cause": ["missing_summary_slide", "missing_volume2_price"],
        "conforming_local_packet_ready": all(item["exists"] for item in packet.values()) and total == 1655000,
    }
    OUT.mkdir(parents=True, exist_ok=True)
    CHECK_JSON.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return data


def build_zip() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    files = [
        (SUMMARY_SLIDE, "SUMMARY_SLIDE/CLARA_Summary_Slide.pdf"),
        (VOLUME_1, "VOLUME_1_TECHNICAL_MANAGEMENT/CLARA_Volume1_Technical.pdf"),
        (PRICE_WORKBOOK, "VOLUME_2_PRICE/CLARA_Volume2_Price_Workbook.xlsx"),
        (VOLUME_3, "VOLUME_3_ADMINISTRATIVE/CLARA_Volume3_Administrative.pdf"),
        (CHECK_JSON, "CHECKS/submission_data_check.json"),
        (CURE / "CLARA_RECONSIDERATION_EMAIL_DRAFT.md", "COVER/CLARA_RECONSIDERATION_EMAIL_DRAFT.md"),
        (CURE / "CORRECTED_PACKET_MANIFEST.md", "CHECKS/CORRECTED_PACKET_MANIFEST.md"),
    ]
    with zipfile.ZipFile(PACKAGE_ZIP, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for src, arcname in files:
            if src.exists():
                zf.write(src, arcname)


def main() -> None:
    build_summary_slide()
    build_price_workbook()
    data = build_check()
    build_zip()
    print(json.dumps({
        "summary_slide": str(SUMMARY_SLIDE),
        "price_workbook": str(PRICE_WORKBOOK),
        "check_json": str(CHECK_JSON),
        "zip": str(PACKAGE_ZIP),
        "zip_bytes": PACKAGE_ZIP.stat().st_size,
        "conforming_local_packet_ready": data["conforming_local_packet_ready"],
        "credential_status": data["credential_status"],
    }, indent=2))


if __name__ == "__main__":
    main()

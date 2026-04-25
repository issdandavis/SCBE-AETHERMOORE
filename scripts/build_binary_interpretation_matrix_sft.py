#!/usr/bin/env python3
"""Convert the binary interpretation workbook into SCBE coding-agent SFT records."""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path(r"C:\Users\issda\Downloads\binary_interpretation_matrix.xlsx")
DEFAULT_OUTPUT = REPO_ROOT / "training-data" / "sft" / "binary_interpretation_matrix_v1.sft.jsonl"
DEFAULT_MANIFEST = REPO_ROOT / "training-data" / "sft" / "binary_interpretation_matrix_v1_manifest.json"

SYSTEM_PROMPT = (
    "You are an SCBE-AETHERMOORE GeoSeal coding agent. Treat binary, numeric encodings, "
    "bit operations, memory layout, and text encodings as deterministic substrate for code generation, "
    "debugging, route execution, and cross-language translation."
)


@dataclass(frozen=True)
class SheetRecord:
    sheet: str
    title: str
    headers: list[str]
    values: dict[str, str]


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _row_values(row: tuple[Any, ...]) -> list[str]:
    return [_clean(value) for value in row]


def _first_nonempty(values: list[str]) -> str:
    for value in values:
        if value:
            return value
    return ""


def _find_header_row(rows: list[list[str]]) -> int:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:12]):
        nonempty = [value for value in row if value]
        alpha = [value for value in nonempty if any(ch.isalpha() for ch in value)]
        score = len(nonempty) + len(alpha)
        if score > best_score and len(nonempty) >= 2:
            best_idx = idx
            best_score = score
    return best_idx


def load_sheet_records(workbook_path: Path) -> list[SheetRecord]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    records: list[SheetRecord] = []
    for ws in wb.worksheets:
        rows = [_row_values(row) for row in ws.iter_rows(values_only=True)]
        nonempty_rows = [row for row in rows if any(row)]
        if len(nonempty_rows) < 3:
            continue
        title = _first_nonempty(nonempty_rows[0]) or ws.title
        header_idx = _find_header_row(nonempty_rows)
        headers = [header or f"column_{idx}" for idx, header in enumerate(nonempty_rows[header_idx], start=1)]
        headers = [header for header in headers if header]
        for row in nonempty_rows[header_idx + 1 :]:
            values = {}
            for idx, header in enumerate(headers):
                if idx >= len(row):
                    continue
                value = row[idx]
                if value:
                    values[header] = value
            if len(values) >= 2:
                records.append(SheetRecord(sheet=ws.title, title=title, headers=headers, values=values))
    return records


def _compact_values(values: dict[str, str]) -> str:
    return "; ".join(f"{key}: {value}" for key, value in values.items())


def record_to_sft(record: SheetRecord, idx: int) -> dict[str, Any]:
    content = _compact_values(record.values)
    instruction = (
        f"Use the Binary Interpretation Matrix sheet '{record.sheet}' to explain this row as deterministic "
        f"coding substrate. Then state how a GeoSeal coding agent should use it."
    )
    assistant = (
        f"Source sheet: {record.sheet}\n"
        f"Topic: {record.title}\n"
        f"Facts: {content}\n\n"
        "GeoSeal use: preserve the exact binary or numeric interpretation before translating code, "
        "choosing a runtime type, decoding text, routing a packet, or explaining low-level behavior."
    )
    return {
        "id": f"binary_interpretation_matrix_v1_{idx:04d}",
        "source": "binary_interpretation_matrix.xlsx",
        "track": "geoseal_coding_binary_substrate",
        "source_type": "workbook_row",
        "quality": "reference",
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": f"{instruction}\n\nRow facts: {content}"},
            {"role": "assistant", "content": assistant},
        ],
        "metadata": {
            "sheet": record.sheet,
            "title": record.title,
            "headers": record.headers,
        },
    }


def build_dataset(workbook_path: Path, output_path: Path, manifest_path: Path) -> dict[str, Any]:
    records = load_sheet_records(workbook_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="\n") as handle:
        for idx, record in enumerate(records, start=1):
            handle.write(json.dumps(record_to_sft(record, idx), ensure_ascii=False) + "\n")
    sheet_counts: dict[str, int] = {}
    for record in records:
        sheet_counts[record.sheet] = sheet_counts.get(record.sheet, 0) + 1
    manifest = {
        "schema_version": "binary_interpretation_matrix_sft_manifest_v1",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "workbook_path": str(workbook_path),
        "output_path": str(output_path),
        "record_count": len(records),
        "sheet_counts": sheet_counts,
        "training_use": [
            "binary-first coding substrate",
            "numeric representation grounding",
            "GeoSeal code generation and route execution explanations",
            "cross-language low-level semantics",
        ],
    }
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=True), encoding="utf-8")
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")
    manifest = build_dataset(args.workbook, args.output, args.manifest)
    print(json.dumps(manifest, indent=2, ensure_ascii=True) if args.json else f"wrote {manifest['record_count']} records")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

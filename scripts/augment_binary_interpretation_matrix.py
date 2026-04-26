#!/usr/bin/env python3
"""Add SCBE extension sheets to the Binary Interpretation Matrix workbook."""

from __future__ import annotations

import argparse
import json
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path(r"C:\Users\issda\Downloads\binary_interpretation_matrix.xlsx")
DEFAULT_REPORT = REPO_ROOT / "artifacts" / "binary_interpretation_matrix" / "augment_report.json"

SCBE_EXTENSION_SHEETS = [
    "Periodic Table Mapping",
    "Information Theory",
    "Error Correction Codes",
    "Feature Space Encoding",
    "Quantum Computing Basics",
]


PERIODIC_64 = [
    (1, "H", "Hydrogen", 1, 1, "nonmetal", 2.20),
    (2, "He", "Helium", 1, 18, "noble gas", None),
    (3, "Li", "Lithium", 2, 1, "alkali metal", 0.98),
    (4, "Be", "Beryllium", 2, 2, "alkaline earth metal", 1.57),
    (5, "B", "Boron", 2, 13, "metalloid", 2.04),
    (6, "C", "Carbon", 2, 14, "nonmetal", 2.55),
    (7, "N", "Nitrogen", 2, 15, "nonmetal", 3.04),
    (8, "O", "Oxygen", 2, 16, "nonmetal", 3.44),
    (9, "F", "Fluorine", 2, 17, "halogen", 3.98),
    (10, "Ne", "Neon", 2, 18, "noble gas", None),
    (11, "Na", "Sodium", 3, 1, "alkali metal", 0.93),
    (12, "Mg", "Magnesium", 3, 2, "alkaline earth metal", 1.31),
    (13, "Al", "Aluminium", 3, 13, "post-transition metal", 1.61),
    (14, "Si", "Silicon", 3, 14, "metalloid", 1.90),
    (15, "P", "Phosphorus", 3, 15, "nonmetal", 2.19),
    (16, "S", "Sulfur", 3, 16, "nonmetal", 2.58),
    (17, "Cl", "Chlorine", 3, 17, "halogen", 3.16),
    (18, "Ar", "Argon", 3, 18, "noble gas", None),
    (19, "K", "Potassium", 4, 1, "alkali metal", 0.82),
    (20, "Ca", "Calcium", 4, 2, "alkaline earth metal", 1.00),
    (21, "Sc", "Scandium", 4, 3, "transition metal", 1.36),
    (22, "Ti", "Titanium", 4, 4, "transition metal", 1.54),
    (23, "V", "Vanadium", 4, 5, "transition metal", 1.63),
    (24, "Cr", "Chromium", 4, 6, "transition metal", 1.66),
    (25, "Mn", "Manganese", 4, 7, "transition metal", 1.55),
    (26, "Fe", "Iron", 4, 8, "transition metal", 1.83),
    (27, "Co", "Cobalt", 4, 9, "transition metal", 1.88),
    (28, "Ni", "Nickel", 4, 10, "transition metal", 1.91),
    (29, "Cu", "Copper", 4, 11, "transition metal", 1.90),
    (30, "Zn", "Zinc", 4, 12, "transition metal", 1.65),
    (31, "Ga", "Gallium", 4, 13, "post-transition metal", 1.81),
    (32, "Ge", "Germanium", 4, 14, "metalloid", 2.01),
    (33, "As", "Arsenic", 4, 15, "metalloid", 2.18),
    (34, "Se", "Selenium", 4, 16, "nonmetal", 2.55),
    (35, "Br", "Bromine", 4, 17, "halogen", 2.96),
    (36, "Kr", "Krypton", 4, 18, "noble gas", 3.00),
    (37, "Rb", "Rubidium", 5, 1, "alkali metal", 0.82),
    (38, "Sr", "Strontium", 5, 2, "alkaline earth metal", 0.95),
    (39, "Y", "Yttrium", 5, 3, "transition metal", 1.22),
    (40, "Zr", "Zirconium", 5, 4, "transition metal", 1.33),
    (41, "Nb", "Niobium", 5, 5, "transition metal", 1.60),
    (42, "Mo", "Molybdenum", 5, 6, "transition metal", 2.16),
    (43, "Tc", "Technetium", 5, 7, "transition metal", 1.90),
    (44, "Ru", "Ruthenium", 5, 8, "transition metal", 2.20),
    (45, "Rh", "Rhodium", 5, 9, "transition metal", 2.28),
    (46, "Pd", "Palladium", 5, 10, "transition metal", 2.20),
    (47, "Ag", "Silver", 5, 11, "transition metal", 1.93),
    (48, "Cd", "Cadmium", 5, 12, "transition metal", 1.69),
    (49, "In", "Indium", 5, 13, "post-transition metal", 1.78),
    (50, "Sn", "Tin", 5, 14, "post-transition metal", 1.96),
    (51, "Sb", "Antimony", 5, 15, "metalloid", 2.05),
    (52, "Te", "Tellurium", 5, 16, "metalloid", 2.10),
    (53, "I", "Iodine", 5, 17, "halogen", 2.66),
    (54, "Xe", "Xenon", 5, 18, "noble gas", 2.60),
    (55, "Cs", "Caesium", 6, 1, "alkali metal", 0.79),
    (56, "Ba", "Barium", 6, 2, "alkaline earth metal", 0.89),
    (57, "La", "Lanthanum", 6, 3, "lanthanide", 1.10),
    (58, "Ce", "Cerium", 6, 3, "lanthanide", 1.12),
    (59, "Pr", "Praseodymium", 6, 3, "lanthanide", 1.13),
    (60, "Nd", "Neodymium", 6, 3, "lanthanide", 1.14),
    (61, "Pm", "Promethium", 6, 3, "lanthanide", 1.13),
    (62, "Sm", "Samarium", 6, 3, "lanthanide", 1.17),
    (63, "Eu", "Europium", 6, 3, "lanthanide", 1.20),
    (64, "Gd", "Gadolinium", 6, 3, "lanthanide", 1.20),
]


def _byte_to_element_index(byte_value: int) -> int:
    """Prime-stride byte projection into the 64-row workbook reference table."""

    return ((byte_value * 47) % 64) + 1


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 12), 46)


def _replace_sheet(wb, title: str):
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def _append_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_sheet(ws)


def _periodic_rows() -> list[list[Any]]:
    rows = []
    by_z = {row[0]: row for row in PERIODIC_64}
    for bucket in range(64):
        start = bucket * 4
        end = start + 3
        element_idx = _byte_to_element_index(start)
        z, symbol, name, period, group, category, electronegativity = by_z[element_idx]
        rows.append(
            [
                bucket,
                f"{start:02X}-{end:02X}",
                f"{start:08b}-{end:08b}",
                element_idx,
                z,
                symbol,
                name,
                period,
                group,
                category,
                "" if electronegativity is None else electronegativity,
                "byte_to_element_index = ((byte * 47) mod 64) + 1; workbook reference table is 64-row bounded",
            ]
        )
    return rows


def augment_workbook(workbook_path: Path, *, backup: bool = True) -> dict[str, Any]:
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    backup_path = None
    if backup:
        stamp = datetime.now().strftime("%Y%m%dT%H%M%S")
        backup_path = workbook_path.with_name(f"{workbook_path.stem}.backup_{stamp}{workbook_path.suffix}")
        shutil.copy2(workbook_path, backup_path)

    wb = load_workbook(workbook_path)

    ws = _replace_sheet(wb, "Periodic Table Mapping")
    _append_rows(
        ws,
        [
            "Bucket",
            "Byte Range Hex",
            "Byte Range Binary",
            "Mapped Element Index",
            "Atomic Number",
            "Symbol",
            "Element",
            "Period",
            "Group",
            "Category",
            "Electronegativity",
            "SCBE Use",
        ],
        _periodic_rows(),
    )

    ws = _replace_sheet(wb, "Information Theory")
    _append_rows(
        ws,
        ["Topic", "Formula / Pattern", "Example", "SCBE Use"],
        [
            ["Self-information", "I(x) = -log2(p(x))", "p=1/256 => 8 bits", "Byte rarity and token surprise scoring"],
            ["Shannon entropy", "H(X) = -sum p(x) log2 p(x)", "Uniform byte distribution => 8 bits", "Tokenizer compression and lane diversity"],
            ["Cross entropy", "H(P,Q) = -sum P(x) log2 Q(x)", "Compare expected vs observed lane tokens", "Training drift / corpus mismatch"],
            ["KL divergence", "D_KL(P||Q)=sum P log2(P/Q)", "Measures distribution shift", "Null-control and benchmark separation"],
            ["Mutual information", "I(X;Y)=H(X)-H(X|Y)", "Concept signal surviving translation", "Cross-primary code concept recovery"],
            ["Hamming distance", "count differing bit positions", "10101010 vs 11110000 => 4", "Binary error and route-delta checks"],
            ["Levenshtein distance", "edit operations between strings", "token vs decoded token", "Text/token transport validation"],
            ["Kolmogorov proxy", "compressed_length(x)", "gzip bytes as rough MDL", "Workflow/tree complexity proxy"],
        ],
    )

    ws = _replace_sheet(wb, "Error Correction Codes")
    _append_rows(
        ws,
        ["Code / Method", "Core Idea", "Example", "SCBE Use"],
        [
            ["Parity bit", "Add one bit to detect odd bit flips", "1011001 + parity", "Cheap lane-integrity signal"],
            ["Checksum", "Sum or hash chunks", "mod-256 byte sum", "Fast packet sanity check"],
            ["CRC-8", "Polynomial remainder over GF(2)", "x^8+x^2+x+1", "Small packet corruption detection"],
            ["CRC-32", "32-bit polynomial checksum", "Ethernet / ZIP", "Artifact and bus-event guard"],
            ["Hamming(7,4)", "4 data bits + 3 parity bits", "single-bit correction", "Tokenizer bit-lane teaching example"],
            ["Hamming(32,26)", "26 data bits + 6 parity bits", "single-error correction", "Longer packed route fields"],
            ["Reed-Solomon", "Symbol-level correction over finite fields", "QR / storage media", "Burst-error and packet recovery analogy"],
            ["DNA degeneracy", "multiple codons map to one amino acid", "64 codons -> 20 amino acids", "Biological redundancy bridge"],
        ],
    )

    ws = _replace_sheet(wb, "Feature Space Encoding")
    _append_rows(
        ws,
        ["Feature", "Extraction", "Range / Type", "SCBE Use"],
        [
            ["byte_value", "int(byte)", "0..255", "Raw substrate anchor"],
            ["high_nibble", "byte >> 4", "0..15", "Hex structure"],
            ["low_nibble", "byte & 0x0F", "0..15", "Hex structure"],
            ["popcount", "count 1 bits", "0..8", "Binary density"],
            ["parity", "popcount mod 2", "0/1", "Error and rhythm feature"],
            ["is_prime_byte", "byte is prime", "bool", "Number-theory marker"],
            ["is_power_of_two", "byte in 2^n", "bool", "Alignment / memory boundary"],
            ["is_fibonacci", "byte in Fibonacci set", "bool", "Sequence marker"],
            ["ascii_class", "control/digit/alpha/punct/extended", "category", "Code/prose separation"],
            ["utf8_role", "ascii/lead/continuation/invalid", "category", "Text encoding safety"],
            ["element_symbol", "prime-stride mapped element", "category", "Structured chemistry lane label"],
            ["element_period", "periodic row", "1..7", "Coarse element geometry"],
            ["element_group", "periodic group", "1..18", "Valence-like grouping"],
            ["electronegativity", "mapped element property", "float/blank", "Experimental affinity lane"],
            ["eml_tree_role", "leaf/operator/boundary", "category", "EML/T experimental symbolic lane"],
        ],
    )

    ws = _replace_sheet(wb, "Quantum Computing Basics")
    _append_rows(
        ws,
        ["Topic", "Representation", "Example", "SCBE Use"],
        [
            ["Classical bit", "0 or 1", "byte = 8 bits", "Binary substrate"],
            ["Qubit", "alpha|0> + beta|1>", "|alpha|^2 + |beta|^2 = 1", "Contrast with probabilistic tokenizer states"],
            ["Bloch sphere", "3D unit-sphere state view", "|+> on equator", "State-geometry analogy with strict boundary"],
            ["Pauli-X", "[[0,1],[1,0]]", "bit flip", "Gate table teaching record"],
            ["Pauli-Z", "[[1,0],[0,-1]]", "phase flip", "Phase/sign lane analogy"],
            ["Hadamard", "1/sqrt(2)[[1,1],[1,-1]]", "creates superposition", "Do not confuse with deterministic transport"],
            ["CNOT", "two-qubit controlled flip", "entangling gate", "Dependency/control-flow analogy"],
            ["Surface code", "2D lattice error correction", "stabilizer checks", "Fault tolerance and bus verification analogy"],
        ],
    )

    wb.save(workbook_path)
    reloaded = load_workbook(workbook_path, read_only=True, data_only=False)
    return {
        "schema_version": "binary_interpretation_matrix_augment_report_v1",
        "generated_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "workbook_path": str(workbook_path),
        "backup_path": str(backup_path) if backup_path else None,
        "sheet_count": len(reloaded.sheetnames),
        "extension_sheets": {name: name in reloaded.sheetnames for name in SCBE_EXTENSION_SHEETS},
        "openxml_load_check": "PASSED",
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--no-backup", action="store_true")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    report = augment_workbook(args.workbook, backup=not args.no_backup)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2, ensure_ascii=True) if args.json else f"augmented {report['sheet_count']} sheets")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


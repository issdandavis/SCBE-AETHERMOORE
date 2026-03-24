"""Build TY2025 Tax Workbook (Excel) from expense CSV and return data."""

import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
money_fmt = "#,##0.00"
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)

# ===== Sheet 1: Expense Detail =====
ws = wb.active
ws.title = "Schedule C Expenses"

ws.merge_cells("A1:F1")
ws["A1"] = "SCHEDULE C BUSINESS EXPENSES - TY2025 - Issac Davis"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A2:F2")
ws["A2"] = "SCBE-AETHERMOORE AI Development | NAICS 541511"
ws["A2"].font = Font(size=11, italic=True)

headers = ["Date", "Vendor", "Amount", "Category", "Sched C Line", "Type"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

csv_path = "artifacts/tax/TY2025_Schedule_C_Expenses.csv"
with open(csv_path, "r") as f:
    reader = csv.DictReader(f)
    row_num = 5
    total = 0
    for r in reader:
        amt = float(r["Amount"])
        ws.cell(row=row_num, column=1, value=r["Date"]).border = thin_border
        ws.cell(row=row_num, column=2, value=r["Vendor"]).border = thin_border
        c = ws.cell(row=row_num, column=3, value=abs(amt))
        c.number_format = money_fmt
        c.border = thin_border
        if r["Type"] == "Refund":
            c.font = Font(color="FF0000")
        ws.cell(row=row_num, column=4, value=r["Category"]).border = thin_border
        ws.cell(row=row_num, column=5, value=r["Schedule_C_Line"]).border = thin_border
        ws.cell(row=row_num, column=6, value=r["Type"]).border = thin_border
        if r["Type"] == "Expense":
            total += abs(amt)
        else:
            total -= abs(amt)
        row_num += 1

row_num += 1
ws.cell(row=row_num, column=2, value="TOTAL DEDUCTIBLE").font = Font(bold=True, size=12)
c = ws.cell(row=row_num, column=3, value=total)
c.number_format = money_fmt
c.font = Font(bold=True, size=12, color="2F5496")

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 10

# ===== Sheet 2: Tax Summary =====
ws2 = wb.create_sheet("Tax Summary")
ws2.merge_cells("A1:D1")
ws2["A1"] = "FEDERAL TAX RETURN SUMMARY - TY2025"
ws2["A1"].font = Font(bold=True, size=14)
ws2.merge_cells("A2:D2")
ws2["A2"] = "Issac Davis | Single | Port Angeles, WA 98362"
ws2["A2"].font = Font(size=11, italic=True)

summary_data = [
    ("", "INCOME", None, ""),
    ("Line 1", "W-2 Wages", 32357.11, "Burger Management Systems"),
    ("Line 3b", "Ordinary Dividends", 145.27, "Cash App Investing / DriveWealth"),
    ("Line 8", "Schedule C Net Loss", -3820.88, "SCBE-AETHERMOORE (AI Dev)"),
    ("Line 8", "Short-Term Capital Gains", 484.49, "Stock sales (Form 8949 Part I)"),
    ("Line 8", "Long-Term Capital Gains", 833.62, "Stock sales (Form 8949 Part II) - 0% rate"),
    ("Line 9", "TOTAL INCOME", 29999.61, ""),
    ("", "", None, ""),
    ("", "DEDUCTIONS", None, ""),
    ("Line 12", "Standard Deduction", -15000.00, ""),
    ("Line 15", "TAXABLE INCOME", 14999.61, ""),
    ("", "", None, ""),
    ("", "TAX", None, ""),
    ("Line 16", "Federal Income Tax", 1561.45, "10% on $11,925 + 12% on $3,075"),
    ("", "Foreign Tax Credit", -0.87, "1099-DIV Box 7"),
    ("Line 24", "TOTAL TAX", 1560.58, ""),
    ("", "", None, ""),
    ("", "PAYMENTS", None, ""),
    ("Line 25a", "Federal Withheld (W-2)", 2684.48, ""),
    ("Line 33", "TOTAL PAYMENTS", 2684.48, ""),
    ("", "", None, ""),
    ("", "RESULT", None, ""),
    ("Line 34", "REFUND", 1123.90, "Direct deposit to Cash App"),
]

section_labels = {"INCOME", "DEDUCTIONS", "TAX", "PAYMENTS", "RESULT"}
bold_labels = {"TOTAL INCOME", "TAXABLE INCOME", "TOTAL TAX", "TOTAL PAYMENTS", "REFUND"}

row = 4
for line, desc, amt, note in summary_data:
    ws2.cell(row=row, column=1, value=line)
    c2 = ws2.cell(row=row, column=2, value=desc)
    if desc in section_labels:
        c2.font = Font(bold=True, size=12, color="2F5496")
    if amt is not None:
        c3 = ws2.cell(row=row, column=3, value=amt)
        c3.number_format = money_fmt
        if desc in bold_labels:
            c3.font = Font(bold=True)
        if desc == "REFUND":
            c3.font = Font(bold=True, size=14, color="008000")
            c2.font = Font(bold=True, size=14, color="008000")
    ws2.cell(row=row, column=4, value=note).font = Font(color="666666", size=9)
    row += 1

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 15
ws2.column_dimensions["D"].width = 45

# ===== Sheet 3: Vendor Summary =====
ws3 = wb.create_sheet("By Vendor")
ws3["A1"] = "EXPENSES BY VENDOR"
ws3["A1"].font = Font(bold=True, size=14)

vendors = {}
with open(csv_path, "r") as f:
    reader = csv.DictReader(f)
    for r in reader:
        v = r["Vendor"].upper().split("*")[0].split(" SUBSCRIPTION")[0].strip()
        if "CLAUDE" in v or "ANTHROPIC" in v:
            v = "Claude / Anthropic"
        elif "OPENAI" in v or "CHATGPT" in v:
            v = "OpenAI / ChatGPT"
        elif "GITHUB" in v:
            v = "GitHub"
        elif "CANVA" in v:
            v = "Canva"
        elif "REPLIT" in v:
            v = "Replit"
        elif "PERPLEXITY" in v:
            v = "Perplexity AI"
        elif "XAI" in v:
            v = "xAI / Grok"
        elif "ZAPIER" in v:
            v = "Zapier"
        elif "NOTION" in v:
            v = "Notion"
        elif "SLACK" in v:
            v = "Slack"
        elif "STRIPE" in v:
            v = "Stripe"
        elif "VERCEL" in v:
            v = "Vercel"
        elif "SHOPIFY" in v:
            v = "Shopify"
        elif "GOOGLE" in v and "ONE" in r["Vendor"].upper():
            v = "Google One"
        else:
            v = r["Vendor"][:30]

        amt = float(r["Amount"])
        mult = 1 if r["Type"] == "Expense" else -1
        vendors[v] = vendors.get(v, 0) + abs(amt) * mult

for col, h in enumerate(["Vendor", "Total", "Schedule C Line"], 1):
    c = ws3.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill

row = 4
for v, amt in sorted(vendors.items(), key=lambda x: -x[1]):
    ws3.cell(row=row, column=1, value=v)
    c = ws3.cell(row=row, column=2, value=amt)
    c.number_format = money_fmt
    ws3.cell(row=row, column=3, value="27a")
    row += 1

row += 1
ws3.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
c = ws3.cell(row=row, column=2, value=sum(vendors.values()))
c.number_format = money_fmt
c.font = Font(bold=True, size=12)

ws3.column_dimensions["A"].width = 25
ws3.column_dimensions["B"].width = 15
ws3.column_dimensions["C"].width = 15

# Save
outpath = "artifacts/tax/TY2025_Issac_Davis_Tax_Workbook.xlsx"
wb.save(outpath)
print(f"Excel workbook saved: {outpath}")
print("3 sheets: Schedule C Expenses, Tax Summary, By Vendor")
